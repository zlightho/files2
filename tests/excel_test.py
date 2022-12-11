import unittest
import os

from openpyxl import Workbook, load_workbook

from excel_function.excel_func import excel_random_numbers_table

TESTDATA_FILEPATH = os.path.join(os.path.dirname(__file__), "testdata.xlsx")
TESTDATA_SHEET_TITLE_NAME = "testdata"


class ExcelTests(unittest.TestCase):
    def setUp(self):
        wb = Workbook()
        ws = wb.create_sheet(title=TESTDATA_SHEET_TITLE_NAME)
        for i in range(1, 10):
            for j in range(1, 10):
                # Заполняем лист случайными строками
                ws.cell(column=i, row=j, value="test")
        wb.save(TESTDATA_FILEPATH)

    def test_before_and_after(self):
        wb = load_workbook(filename=TESTDATA_FILEPATH)
        ws2 = wb[TESTDATA_SHEET_TITLE_NAME]
        for i in range(1, ws2.max_column):
            for j in range(1, ws2.max_row):
                self.assertEqual(isinstance(ws2[i][j].value, str), True)
        excel_random_numbers_table(TESTDATA_SHEET_TITLE_NAME, 10, 10)
        for i in range(1, ws2.max_column):
            for j in range(1, ws2.max_row):
                self.assertEqual(isinstance(ws2[i][j].value, int), True)
